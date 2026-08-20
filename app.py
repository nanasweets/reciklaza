import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import io
import re
import plotly.express as px
import plotly.graph_objects as go
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import os
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')

# ============================
# REGISTRACIJA FONTA
# ============================

FONT_NAME = 'Helvetica'
try:
    pdfmetrics.registerFont(TTFont('DejaVu', '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf'))
    pdfmetrics.registerFont(TTFont('DejaVu-Bold', '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf'))
    FONT_NAME = 'DejaVu'
except:
    try:
        pdfmetrics.registerFont(TTFont('DejaVu', 'DejaVuSans.ttf'))
        pdfmetrics.registerFont(TTFont('DejaVu-Bold', 'DejaVuSans-Bold.ttf'))
        FONT_NAME = 'DejaVu'
    except:
        FONT_NAME = 'Helvetica'

# ============================
# KONFIGURACIJA
# ============================

st.set_page_config(page_title="Izveštaj o reciklaži", layout="wide")

SHEET_RECIKLAZA = "reciklaža"
SHEET_REKLAMACIJE = "reklamacije ukupno"

MAPPING = {
    "reklamacija otvorena od strane": "Reklamacija: Created By",
    "model reklamacije": "RMA Model",
    "klasifikacija": "Klasifikacija",
    "način razduženja kupca": "Način razduženja kupca",
    "serijski broj": "Serijski broj uređaja",
    "datum otvaranja": "Reklamacija: Created Date",
}

KOLONE_OUTPUT = [
    "Broj reklamacije", "ID uređaja", "Naziv artikla", "Brend",
    "Robna grupa", "Količina", "Klasifikacija reciklaže", "Datum obrade",
    "Paleta", "Nabavna cena", "ID ulaza", "Skladišna lokacija",
    "Pozicija u rafu", "Klasifikacija štete", "prevoznik",
    "broj pošiljke", "vrednost fakture",
    "reklamacija otvorena od strane", "model reklamacije",
    "klasifikacija", "način razduženja kupca", "serijski broj",
    "datum otvaranja",
    "Starost (dani)",
    "Rok (30 dana)",
    "Prekoračenje (dani)",
    "Status roka",
]

HEADER_BG = "1F4E79"
HEADER_FONT = "FFFFFF"
LOOKUP_BG = "D9E1F2"
LOOKUP_FONT = "1F4E79"

# Parametri za prodaju trećem licu
if "kurs_evra" not in st.session_state:
    st.session_state.kurs_evra = 117.0
if "komada_po_paleti" not in st.session_state:
    st.session_state.komada_po_paleti = 50
if "cena_po_paleti_evri" not in st.session_state:
    st.session_state.cena_po_paleti_evri = 110.0

# ============================
# FUNKCIJE ZA ČIŠĆENJE
# ============================

def ocisti_datum(vrednost):
    if pd.isna(vrednost):
        return pd.NaT
    if isinstance(vrednost, (pd.Timestamp, datetime)):
        return vrednost
    tekst = str(vrednost).strip()
    try:
        return pd.to_datetime(tekst, format='%d.%m.%Y', errors='coerce')
    except:
        pass
    try:
        if ' 0:00:00' in tekst:
            tekst = tekst.replace(' 0:00:00', '')
        return pd.to_datetime(tekst, format='%Y-%m-%d', errors='coerce')
    except:
        pass
    try:
        return pd.to_datetime(tekst, errors='coerce')
    except:
        return pd.NaT

def cisti_broj(vrednost):
    if pd.isna(vrednost):
        return ""
    tekst = str(vrednost).strip()
    cifre = re.sub(r'[^0-9]', '', tekst)
    return cifre

def pronadji_kolonu_broja(df):
    for kol in df.columns:
        if "broj" in kol.lower() and "reklamacije" in kol.lower():
            return kol
        if "reklamacija" in kol.lower() and ("broj" in kol.lower() or "id" in kol.lower()):
            return kol
    return df.columns[0]

# ============================
# OBRADA PODATAKA SA ROKOVIMA
# ============================

def ucitaj_podatke(fajl):
    xl = pd.ExcelFile(fajl)
    if SHEET_RECIKLAZA not in xl.sheet_names:
        st.error(f"Sheet '{SHEET_RECIKLAZA}' nije pronađen!")
        return None, None
    if SHEET_REKLAMACIJE not in xl.sheet_names:
        st.error(f"Sheet '{SHEET_REKLAMACIJE}' nije pronađen!")
        return None, None
    reciklaza = pd.read_excel(fajl, sheet_name=SHEET_RECIKLAZA, header=0)
    reklamacije = pd.read_excel(fajl, sheet_name=SHEET_REKLAMACIJE, header=0)
    
    if "Datum obrade" in reciklaza.columns:
        reciklaza["Datum obrade"] = reciklaza["Datum obrade"].apply(ocisti_datum)
        reciklaza = reciklaza[reciklaza["Datum obrade"].notna()]
    
    return reciklaza, reklamacije

def pripremi_lookup(reklamacije):
    kolona_broja = pronadji_kolonu_broja(reklamacije)
    rek = reklamacije.copy()
    rek["_kljuc"] = rek[kolona_broja].apply(cisti_broj)
    rek = rek[rek["_kljuc"] != ""]
    return rek.set_index("_kljuc")

def izracunaj_rokove(df):
    df_copy = df.copy()
    
    if "datum otvaranja" in df_copy.columns and "Datum obrade" in df_copy.columns:
        df_copy["datum otvaranja"] = pd.to_datetime(df_copy["datum otvaranja"], errors="coerce")
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        
        danas = datetime.now()
        df_copy = df_copy[
            (df_copy["Datum obrade"].notna()) & 
            (df_copy["Datum obrade"] <= danas)
        ]
        
        if len(df_copy) > 0:
            df_copy["Starost (dani)"] = (df_copy["Datum obrade"] - df_copy["datum otvaranja"]).dt.days
            df_copy["Rok (30 dana)"] = df_copy["datum otvaranja"] + pd.Timedelta(days=30)
            df_copy["Prekoračenje (dani)"] = df_copy["Starost (dani)"] - 30
            df_copy["Status roka"] = df_copy["Prekoračenje (dani)"].apply(
                lambda x: "✅ U roku" if (pd.isna(x) or x <= 0) else "❌ Prekoračen"
            )
        else:
            df_copy["Starost (dani)"] = None
            df_copy["Rok (30 dana)"] = None
            df_copy["Prekoračenje (dani)"] = None
            df_copy["Status roka"] = "N/A"
    else:
        df_copy["Starost (dani)"] = None
        df_copy["Rok (30 dana)"] = None
        df_copy["Prekoračenje (dani)"] = None
        df_copy["Status roka"] = "Nedostaju podaci"
    
    return df_copy

def spoji(reciklaza, lookup):
    df = reciklaza.copy()
    df["_kljuc"] = df["Broj reklamacije"].apply(cisti_broj)
    for out_kol, src_kol in MAPPING.items():
        if src_kol not in lookup.columns:
            df[out_kol] = None
            continue
        vrednosti = []
        for kljuc in df["_kljuc"]:
            if kljuc == "":
                vrednosti.append(None)
            elif kljuc in lookup.index:
                vrednosti.append(lookup.at[kljuc, src_kol])
            else:
                vrednosti.append(None)
        df[out_kol] = vrednosti
    df = df.drop(columns=["_kljuc"])
    df = izracunaj_rokove(df)
    extras = [c for c in df.columns if c not in KOLONE_OUTPUT]
    finalne = [c for c in KOLONE_OUTPUT if c in df.columns] + extras
    return df[finalne]

# ============================
# PRODAJA TREĆEM LICU
# ============================

def generisi_predlog_prodaje(df, kurs_evra=117, komada_po_paleti=50, cena_po_paleti_evri=110):
    df_copy = df.copy()
    
    if "Starost (dani)" not in df_copy.columns:
        return None, "Nema podataka o starosti reklamacija"
    
    if "Datum obrade" not in df_copy.columns:
        return None, "Nema podataka o datumu prijema"
    
    df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
    df_copy["Starost (dani)"] = pd.to_numeric(df_copy["Starost (dani)"], errors="coerce")
    
    danas = datetime.now()
    df_copy["Vreme u magacinu (dani)"] = (danas - df_copy["Datum obrade"]).dt.days
    
    # FILTER: Starost > 30 dana I Vreme u magacinu > 30 dana (ukupno > 60 dana)
    df_predlog = df_copy[
        (df_copy["Starost (dani)"] > 30) & 
        (df_copy["Vreme u magacinu (dani)"] > 30) &
        (df_copy["Datum obrade"].notna())
    ].copy()
    
    if len(df_predlog) == 0:
        return None, "Nema artikala koji ispunjavaju kriterijume za prodaju (starost >30 dana i vreme u magacinu >30 dana)"
    
    df_predlog["Ukupna nabavna vrednost (RSD)"] = df_predlog["Nabavna cena"] * df_predlog["Količina"]
    df_predlog["Broj paleta"] = (df_predlog["Količina"] / komada_po_paleti).apply(lambda x: int(x) + (1 if x % 1 > 0 else 0))
    df_predlog["Vrednost ponude (€)"] = df_predlog["Broj paleta"] * cena_po_paleti_evri
    df_predlog["Vrednost ponude (RSD)"] = df_predlog["Vrednost ponude (€)"] * kurs_evra
    df_predlog["Gubitak (RSD)"] = df_predlog["Ukupna nabavna vrednost (RSD)"] - df_predlog["Vrednost ponude (RSD)"]
    df_predlog["Ukupna starost (dani)"] = df_predlog["Starost (dani)"] + df_predlog["Vreme u magacinu (dani)"]
    
    sazetak = {
        "Ukupno artikala": df_predlog["Količina"].sum(),
        "Ukupno paleta": df_predlog["Broj paleta"].sum(),
        "Ukupna nabavna vrednost (RSD)": df_predlog["Ukupna nabavna vrednost (RSD)"].sum(),
        "Ukupna vrednost ponude (€)": df_predlog["Vrednost ponude (€)"].sum(),
        "Ukupna vrednost ponude (RSD)": df_predlog["Vrednost ponude (RSD)"].sum(),
        "Ukupni gubitak (RSD)": df_predlog["Gubitak (RSD)"].sum(),
        "Broj artikala za prodaju": len(df_predlog),
        "Prosečna starost reklamacije (dani)": df_predlog["Starost (dani)"].mean().round(0),
        "Prosečno vreme u magacinu (dani)": df_predlog["Vreme u magacinu (dani)"].mean().round(0),
        "Prosečna ukupna starost (dani)": df_predlog["Ukupna starost (dani)"].mean().round(0),
    }
    
    df_predlog = df_predlog.sort_values("Gubitak (RSD)", ascending=False)
    return df_predlog, sazetak

# ============================
# POMOĆNE FUNKCIJE ZA EXCEL
# ============================

def excel_compatible(value):
    if pd.isna(value) or value is None:
        return None
    if isinstance(value, (int, float, str, bool)):
        return value
    if isinstance(value, (pd.Timestamp, datetime)):
        return value
    return str(value)

def primeni_stil_header(ws, max_col, lookup_kolone=None):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        kol = ws.cell(row=1, column=col_idx).value
        if lookup_kolone and kol in lookup_kolone:
            cell.fill = PatternFill("solid", fgColor=LOOKUP_BG)
            cell.font = Font(bold=True, color=LOOKUP_FONT, name="Arial", size=10)
        else:
            cell.fill = PatternFill("solid", fgColor=HEADER_BG)
            cell.font = Font(bold=True, color=HEADER_FONT, name="Arial", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[1].height = 36

def primeni_stil_podaci(ws, max_col, max_row, lookup_kolone=None):
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            col_name = ws.cell(row=1, column=col_idx).value
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
            if lookup_kolone and col_name in lookup_kolone:
                cell.fill = PatternFill("solid", fgColor="EEF3FB")
            elif row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor="F5F8FD")

def prilagodi_sirine(ws, df):
    for col_idx, kol in enumerate(df.columns, start=1):
        max_len = len(str(kol))
        for row_idx in range(2, min(len(df) + 2, 52)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

def dodaj_legendu(ws, legenda_row, lookup_kolone=None):
    ws.cell(row=legenda_row, column=1, value="Legenda:").font = Font(bold=True, name="Arial", size=9)
    ws.cell(row=legenda_row + 1, column=1, value="■ Tamno plava = ručno uneti podaci").font = Font(name="Arial", size=9, color=HEADER_BG)
    ws.cell(row=legenda_row + 2, column=1, value="■ Svetlo plava = automatski povučeno").font = Font(name="Arial", size=9, color=LOOKUP_FONT)
    ws.cell(row=legenda_row + 4, column=1, value=f"Generisano: {datetime.now().strftime('%d.%m.%Y %H:%M')}").font = Font(italic=True, name="Arial", size=9, color="888888")

# ============================
# FUNKCIJE ZA UPIS SHEET-OVA
# ============================

def upisi_zbirno(ws, df):
    lookup_kolone = set(MAPPING.keys())
    for col_idx, kol in enumerate(df.columns, start=1):
        ws.cell(row=1, column=col_idx, value=kol)
    primeni_stil_header(ws, len(df.columns), lookup_kolone)
    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx).value = excel_compatible(value)
    primeni_stil_podaci(ws, len(df.columns), len(df) + 1, lookup_kolone)
    prilagodi_sirine(ws, df)
    dodaj_legendu(ws, len(df) + 3, lookup_kolone)
    ws.freeze_panes = "A2"

def upisi_kpi(ws, df):
    ukupno_artikala = df["Količina"].sum() if "Količina" in df.columns else 0
    ukupna_vrednost = (df["Nabavna cena"] * df["Količina"]).sum() if "Nabavna cena" in df.columns and "Količina" in df.columns else 0
    broj_brendova = df["Brend"].nunique() if "Brend" in df.columns else 0
    broj_grupa = df["Robna grupa"].nunique() if "Robna grupa" in df.columns else 0
    
    u_roku = df[df["Status roka"] == "✅ U roku"].shape[0] if "Status roka" in df.columns else 0
    prekoraceno = df[df["Status roka"] == "❌ Prekoračen"].shape[0] if "Status roka" in df.columns else 0
    
    if "Datum obrade" in df.columns:
        df_datum = df.copy()
        df_datum["Datum obrade"] = pd.to_datetime(df_datum["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_datum = df_datum[(df_datum["Datum obrade"].notna()) & (df_datum["Datum obrade"] <= danas)]
        if len(df_datum) > 0:
            min_datum = df_datum["Datum obrade"].min()
            max_datum = df_datum["Datum obrade"].max()
            period = f"{min_datum.strftime('%d.%m.%Y') if pd.notna(min_datum) else 'N/A'} - {max_datum.strftime('%d.%m.%Y') if pd.notna(max_datum) else 'N/A'}"
        else:
            period = "N/A"
    else:
        period = "N/A"
    
    podaci = [
        ["Ključni pokazatelj", "Vrednost"],
        ["Ukupan broj artikala", ukupno_artikala],
        ["Ukupna vrednost (RSD)", f"{ukupna_vrednost:,.2f}"],
        ["Broj brendova", broj_brendova],
        ["Broj robnih grupa", broj_grupa],
        ["Period", period],
        ["Broj redova", len(df)],
        ["✅ U roku (≤30 dana)", u_roku],
        ["❌ Prekoračen (>30 dana)", prekoraceno],
    ]
    
    for row_idx, row in enumerate(podaci, start=1):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    for col_idx in [1, 2]:
        ws.column_dimensions[get_column_letter(col_idx)].width = 30
    
    for row_idx in range(1, len(podaci) + 1):
        ws.cell(row=row_idx, column=1).font = Font(bold=True, name="Arial", size=11)
        ws.cell(row=row_idx, column=2).font = Font(name="Arial", size=11)
    
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color=HEADER_BG)
    ws.cell(row=1, column=2).font = Font(bold=True, size=12, color=HEADER_BG)

def upisi_grupisano(ws, df, kolona, naslov):
    if kolona not in df.columns:
        ws.cell(row=1, column=1, value=f"Kolona '{kolona}' ne postoji")
        return
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    grupisan = df_copy[kolona].value_counts().reset_index()
    grupisan.columns = [kolona, "Broj artikala"]
    
    if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        vrednost_grupisano = df_copy.groupby(kolona)["Ukupna vrednost"].sum().reset_index()
        vrednost_grupisano.columns = [kolona, "Ukupna vrednost (RSD)"]
        grupisan = grupisan.merge(vrednost_grupisano, on=kolona, how="left")
    else:
        grupisan["Ukupna vrednost (RSD)"] = 0
    
    ws.cell(row=1, column=1, value=kolona)
    ws.cell(row=1, column=2, value="Broj artikala")
    ws.cell(row=1, column=3, value="Ukupna vrednost (RSD)")
    primeni_stil_header(ws, 3)
    
    for row_idx, row in enumerate(grupisan.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1).value = excel_compatible(row[0])
        ws.cell(row=row_idx, column=2).value = excel_compatible(row[1])
        if len(row) > 2 and row[2] is not None:
            ws.cell(row=row_idx, column=3).value = excel_compatible(row[2])
    primeni_stil_podaci(ws, 3, len(grupisan) + 1)
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

def upisi_top_brendove(ws, df):
    if "Brend" not in df.columns:
        ws.cell(row=1, column=1, value="Kolona 'Brend' ne postoji")
        return
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        brend_analiza = df_copy.groupby("Brend").agg({
            "Broj reklamacije": "count",
            "Ukupna vrednost": "sum"
        }).reset_index()
        brend_analiza.columns = ["Brend", "Broj artikala", "Ukupna vrednost (RSD)"]
    else:
        brend_analiza = df_copy["Brend"].value_counts().head(10).reset_index()
        brend_analiza.columns = ["Brend", "Broj artikala"]
        brend_analiza["Ukupna vrednost (RSD)"] = 0
    
    brend_analiza = brend_analiza.sort_values("Broj artikala", ascending=False).head(10)
    
    ws.cell(row=1, column=1, value="Top 10 brendova")
    ws.cell(row=1, column=1).font = Font(bold=True, size=12, color=HEADER_BG)
    ws.cell(row=2, column=1, value="Brend")
    ws.cell(row=2, column=2, value="Broj artikala")
    ws.cell(row=2, column=3, value="Ukupna vrednost (RSD)")
    primeni_stil_header(ws, 3)
    
    for idx, row in enumerate(brend_analiza.itertuples(index=False), start=3):
        ws.cell(row=idx, column=1).value = excel_compatible(row[0])
        ws.cell(row=idx, column=2).value = excel_compatible(row[1])
        ws.cell(row=idx, column=3).value = excel_compatible(row[2])
    primeni_stil_podaci(ws, 3, len(brend_analiza) + 2)
    
    for col in [1, 2, 3]:
        ws.column_dimensions[get_column_letter(col)].width = 25

def upisi_top_grupe(ws, df):
    if "Robna grupa" not in df.columns:
        ws.cell(row=1, column=1, value="Kolona 'Robna grupa' ne postoji")
        return
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        grupe = df_copy.groupby("Robna grupa").agg({
            "Broj reklamacije": "count",
            "Ukupna vrednost": "sum"
        }).reset_index()
        grupe.columns = ["Robna grupa", "Broj artikala", "Ukupna vrednost (RSD)"]
    else:
        grupe = df_copy["Robna grupa"].value_counts().head(10).reset_index()
        grupe.columns = ["Robna grupa", "Broj artikala"]
        grupe["Ukupna vrednost (RSD)"] = 0
    
    grupe = grupe.sort_values("Broj artikala", ascending=False).head(10)
    
    ws.cell(row=1, column=1, value="Robna grupa")
    ws.cell(row=1, column=2, value="Broj artikala")
    ws.cell(row=1, column=3, value="Ukupna vrednost (RSD)")
    primeni_stil_header(ws, 3)
    
    for row_idx, row in enumerate(grupe.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1).value = excel_compatible(row[0])
        ws.cell(row=row_idx, column=2).value = excel_compatible(row[1])
        ws.cell(row=row_idx, column=3).value = excel_compatible(row[2])
    primeni_stil_podaci(ws, 3, len(grupe) + 1)
    
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25

def upisi_vremenski_trend(ws, df):
    if "Datum obrade" not in df.columns:
        ws.cell(row=1, column=1, value="Kolona 'Datum obrade' ne postoji")
        return
    
    df_datum = df.copy()
    df_datum["Datum obrade"] = pd.to_datetime(df_datum["Datum obrade"], errors="coerce")
    danas = datetime.now()
    df_datum = df_datum[
        (df_datum["Datum obrade"].notna()) & 
        (df_datum["Datum obrade"] <= danas) &
        (df_datum["Datum obrade"] >= datetime(2020, 1, 1))
    ]
    
    if len(df_datum) == 0:
        ws.cell(row=1, column=1, value="Nema validnih datuma za prikaz")
        return
    
    df_datum["Mesec"] = df_datum["Datum obrade"].dt.to_period("M")
    
    trend = df_datum.groupby("Mesec").agg({
        "Broj reklamacije": "count"
    }).reset_index()
    trend.columns = ["Mesec", "Broj artikala"]
    trend["Mesec"] = trend["Mesec"].astype(str)
    
    if "Nabavna cena" in df_datum.columns and "Količina" in df_datum.columns:
        df_datum["Ukupna vrednost"] = df_datum["Nabavna cena"] * df_datum["Količina"]
        vrednost_po_mesecu = df_datum.groupby("Mesec")["Ukupna vrednost"].sum().reset_index()
        vrednost_po_mesecu.columns = ["Mesec", "Ukupna vrednost (RSD)"]
        vrednost_po_mesecu["Mesec"] = vrednost_po_mesecu["Mesec"].astype(str)
        trend = trend.merge(vrednost_po_mesecu, on="Mesec", how="left")
    
    ws.cell(row=1, column=1, value="Mesec")
    ws.cell(row=1, column=2, value="Broj artikala")
    if "Ukupna vrednost (RSD)" in trend.columns:
        ws.cell(row=1, column=3, value="Ukupna vrednost (RSD)")
        primeni_stil_header(ws, 3)
    else:
        primeni_stil_header(ws, 2)
    
    for row_idx, row in enumerate(trend.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1).value = excel_compatible(row[0])
        ws.cell(row=row_idx, column=2).value = excel_compatible(row[1])
        if len(row) > 2:
            ws.cell(row=row_idx, column=3).value = excel_compatible(row[2])
    
    primeni_stil_podaci(ws, len(trend.columns), len(trend) + 1)
    
    for col_idx in range(1, len(trend.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25

def upisi_prevoznike(ws, df):
    if "prevoznik" not in df.columns:
        ws.cell(row=1, column=1, value="Kolona 'prevoznik' ne postoji")
        return
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        prevoznici = df_copy.groupby("prevoznik").agg({
            "Broj reklamacije": "count",
            "Ukupna vrednost": "sum"
        }).reset_index()
        prevoznici.columns = ["Prevoznik", "Broj artikala", "Ukupna vrednost (RSD)"]
    else:
        prevoznici = df_copy["prevoznik"].value_counts().reset_index()
        prevoznici.columns = ["Prevoznik", "Broj artikala"]
        prevoznici["Ukupna vrednost (RSD)"] = 0
    
    if "Klasifikacija štete" in df_copy.columns:
        stete_prevoz = df_copy.groupby(["prevoznik", "Klasifikacija štete"]).size().reset_index()
        stete_prevoz.columns = ["Prevoznik", "Klasifikacija štete", "Broj"]
    
    ws.cell(row=1, column=1, value="Prevoznik")
    ws.cell(row=1, column=2, value="Broj artikala")
    ws.cell(row=1, column=3, value="Ukupna vrednost (RSD)")
    ws.cell(row=1, column=5, value="Štete po prevozniku")
    ws.cell(row=2, column=5, value="Prevoznik")
    ws.cell(row=2, column=6, value="Klasifikacija štete")
    ws.cell(row=2, column=7, value="Broj")
    
    primeni_stil_header(ws, 3)
    for col in [5, 6, 7]:
        ws.cell(row=1, column=col).font = Font(bold=True, size=11, color=HEADER_BG)
        ws.cell(row=2, column=col).font = Font(bold=True, size=10, color=HEADER_BG)
    
    for row_idx, row in enumerate(prevoznici.itertuples(index=False), start=2):
        ws.cell(row=row_idx, column=1).value = excel_compatible(row[0])
        ws.cell(row=row_idx, column=2).value = excel_compatible(row[1])
        ws.cell(row=row_idx, column=3).value = excel_compatible(row[2])
    primeni_stil_podaci(ws, 3, len(prevoznici) + 1)
    
    if "Klasifikacija štete" in df_copy.columns:
        for row_idx, row in enumerate(stete_prevoz.itertuples(index=False), start=3):
            ws.cell(row=row_idx, column=5).value = excel_compatible(row[0])
            ws.cell(row=row_idx, column=6).value = excel_compatible(row[1])
            ws.cell(row=row_idx, column=7).value = excel_compatible(row[2])
        primeni_stil_podaci(ws, 7, len(stete_prevoz) + 3)
    
    for col in [1, 2, 3, 5, 6, 7]:
        ws.column_dimensions[get_column_letter(col)].width = 25

def upisi_analiza_rokova(ws, df):
    ws.cell(row=1, column=1, value="Analiza rokova reklamacija")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=HEADER_BG)
    
    row = 3
    
    if "Status roka" not in df.columns or "Starost (dani)" not in df.columns:
        ws.cell(row=row, column=1, value="Nedostaju podaci za analizu rokova")
        return
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    if len(df_copy) == 0:
        ws.cell(row=row, column=1, value="Nema validnih podataka za prikaz")
        return
    
    u_roku = df_copy[df_copy["Status roka"] == "✅ U roku"].shape[0]
    prekoraceno = df_copy[df_copy["Status roka"] == "❌ Prekoračen"].shape[0]
    ukupno = len(df_copy)
    
    ws.cell(row=row, column=1, value="Statistika rokova")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupno reklamacija")
    ws.cell(row=row, column=2, value=ukupno)
    row += 1
    
    ws.cell(row=row, column=1, value="✅ U roku (≤30 dana)")
    ws.cell(row=row, column=2, value=u_roku)
    ws.cell(row=row, column=3, value=f"{u_roku/ukupno*100:.1f}%" if ukupno > 0 else "0%")
    row += 1
    
    ws.cell(row=row, column=1, value="❌ Prekoračen (>30 dana)")
    ws.cell(row=row, column=2, value=prekoraceno)
    ws.cell(row=row, column=3, value=f"{prekoraceno/ukupno*100:.1f}%" if ukupno > 0 else "0%")
    row += 2
    
    ws.cell(row=row, column=1, value="Prosečno prekoračenje (dani)")
    ws.cell(row=row, column=2, value=excel_compatible(df_copy["Prekoračenje (dani)"].mean().round(0)))
    row += 1
    
    ws.cell(row=row, column=1, value="Maksimalno prekoračenje (dani)")
    ws.cell(row=row, column=2, value=excel_compatible(df_copy["Prekoračenje (dani)"].max()))
    row += 1
    
    ws.cell(row=row, column=1, value="Prosečna starost (dani)")
    ws.cell(row=row, column=2, value=excel_compatible(df_copy["Starost (dani)"].mean().round(0)))
    row += 2
    
    ws.cell(row=row, column=1, value="Top 10 najvećih prekoračenja roka")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    
    headers = ["Broj reklamacije", "Naziv artikla", "Starost (dani)", "Prekoračenje (dani)", "Otvorio"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=h)
    primeni_stil_header(ws, len(headers))
    row += 1
    
    najgori = df_copy.nlargest(10, "Prekoračenje (dani)")[["Broj reklamacije", "Naziv artikla", "Starost (dani)", "Prekoračenje (dani)", "reklamacija otvorena od strane"]]
    for _, r in najgori.iterrows():
        ws.cell(row=row, column=1).value = excel_compatible(r["Broj reklamacije"])
        ws.cell(row=row, column=2).value = excel_compatible(r["Naziv artikla"])
        ws.cell(row=row, column=3).value = excel_compatible(r["Starost (dani)"])
        ws.cell(row=row, column=4).value = excel_compatible(r["Prekoračenje (dani)"])
        ws.cell(row=row, column=5).value = excel_compatible(r["reklamacija otvorena od strane"])
        row += 1
    
    for col in range(1, 6):
        ws.column_dimensions[get_column_letter(col)].width = 30

def upisi_predlog_prodaje(ws, df, kurs_evra=117, komada_po_paleti=50, cena_po_paleti_evri=110):
    ws.cell(row=1, column=1, value="PREDLOG ZA PRODAJU TREĆEM LICU")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=HEADER_BG)
    
    row = 3
    
    if "Starost (dani)" not in df.columns or "Datum obrade" not in df.columns:
        ws.cell(row=row, column=1, value="Nedostaju podaci za generisanje predloga")
        return
    
    df_copy = df.copy()
    df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
    df_copy["Starost (dani)"] = pd.to_numeric(df_copy["Starost (dani)"], errors="coerce")
    
    danas = datetime.now()
    df_copy["Vreme u magacinu (dani)"] = (danas - df_copy["Datum obrade"]).dt.days
    
    # FILTER: Starost > 30 dana I Vreme u magacinu > 30 dana
    df_predlog = df_copy[
        (df_copy["Starost (dani)"] > 30) & 
        (df_copy["Vreme u magacinu (dani)"] > 30) &
        (df_copy["Datum obrade"].notna())
    ].copy()
    
    if len(df_predlog) == 0:
        ws.cell(row=row, column=1, value="Nema artikala koji ispunjavaju kriterijume za prodaju")
        return
    
    # Parametri
    ws.cell(row=row, column=1, value="PARAMETRI PREDLOGA")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    ws.cell(row=row, column=1, value="Kurs evra (RSD)")
    ws.cell(row=row, column=2, value=kurs_evra)
    row += 1
    
    ws.cell(row=row, column=1, value="Komada po paleti")
    ws.cell(row=row, column=2, value=komada_po_paleti)
    row += 1
    
    ws.cell(row=row, column=1, value="Cena po paleti (€)")
    ws.cell(row=row, column=2, value=cena_po_paleti_evri)
    row += 2
    
    # Kriterijum
    ws.cell(row=row, column=1, value="KRITERIJUM ZA PRODAJU:")
    ws.cell(row=row, column=1).font = Font(bold=True, size=11)
    row += 1
    
    ws.cell(row=row, column=1, value="Starost reklamacije > 30 dana (zakonski rok istekao)")
    row += 1
    
    ws.cell(row=row, column=1, value="Vreme u magacinu > 30 dana (čeka na prodaju)")
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupno > 60 dana od otvaranja reklamacije")
    row += 2
    
    # Izračunavanja
    df_predlog["Ukupna nabavna vrednost (RSD)"] = df_predlog["Nabavna cena"] * df_predlog["Količina"]
    df_predlog["Broj paleta"] = (df_predlog["Količina"] / komada_po_paleti).apply(lambda x: int(x) + (1 if x % 1 > 0 else 0))
    df_predlog["Vrednost ponude (€)"] = df_predlog["Broj paleta"] * cena_po_paleti_evri
    df_predlog["Vrednost ponude (RSD)"] = df_predlog["Vrednost ponude (€)"] * kurs_evra
    df_predlog["Gubitak (RSD)"] = df_predlog["Ukupna nabavna vrednost (RSD)"] - df_predlog["Vrednost ponude (RSD)"]
    df_predlog["Ukupna starost (dani)"] = df_predlog["Starost (dani)"] + df_predlog["Vreme u magacinu (dani)"]
    
    # Header
    headers = [
        "Broj reklamacije", "Naziv artikla", "Brend", "Količina",
        "Nabavna cena (RSD)", "Ukupna nabavna (RSD)",
        "Starost rekl. (dani)", "Vreme u mag. (dani)", "Ukupno (dani)",
        "Broj paleta", "Ponuda (€)", "Ponuda (RSD)", "Gubitak (RSD)"
    ]
    
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=row, column=col_idx, value=h)
    primeni_stil_header(ws, len(headers))
    row += 1
    
    for _, r in df_predlog.iterrows():
        ws.cell(row=row, column=1).value = excel_compatible(r["Broj reklamacije"])
        ws.cell(row=row, column=2).value = excel_compatible(r["Naziv artikla"])
        ws.cell(row=row, column=3).value = excel_compatible(r["Brend"])
        ws.cell(row=row, column=4).value = excel_compatible(r["Količina"])
        ws.cell(row=row, column=5).value = excel_compatible(r["Nabavna cena"])
        ws.cell(row=row, column=6).value = excel_compatible(r["Ukupna nabavna vrednost (RSD)"])
        ws.cell(row=row, column=7).value = excel_compatible(r["Starost (dani)"])
        ws.cell(row=row, column=8).value = excel_compatible(r["Vreme u magacinu (dani)"])
        ws.cell(row=row, column=9).value = excel_compatible(r["Ukupna starost (dani)"])
        ws.cell(row=row, column=10).value = excel_compatible(r["Broj paleta"])
        ws.cell(row=row, column=11).value = excel_compatible(r["Vrednost ponude (€)"])
        ws.cell(row=row, column=12).value = excel_compatible(r["Vrednost ponude (RSD)"])
        ws.cell(row=row, column=13).value = excel_compatible(r["Gubitak (RSD)"])
        row += 1
    
    # Sažetak
    row += 2
    ws.cell(row=row, column=1, value="SAŽETAK PREDLOGA")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupno artikala")
    ws.cell(row=row, column=2, value=df_predlog["Količina"].sum())
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupno paleta")
    ws.cell(row=row, column=2, value=df_predlog["Broj paleta"].sum())
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupna nabavna vrednost (RSD)")
    ws.cell(row=row, column=2, value=f"{df_predlog['Ukupna nabavna vrednost (RSD)'].sum():,.2f}")
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupna vrednost ponude (€)")
    ws.cell(row=row, column=2, value=f"{df_predlog['Vrednost ponude (€)'].sum():,.2f} €")
    row += 1
    
    ws.cell(row=row, column=1, value="Ukupna vrednost ponude (RSD)")
    ws.cell(row=row, column=2, value=f"{df_predlog['Vrednost ponude (RSD)'].sum():,.2f} RSD")
    row += 1
    
    ws.cell(row=row, column=1, value="UKUPAN GUBITAK (RSD)")
    ws.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
    ws.cell(row=row, column=2, value=f"{df_predlog['Gubitak (RSD)'].sum():,.2f} RSD")
    ws.cell(row=row, column=2).font = Font(bold=True, color="FF0000")
    row += 1
    
    ws.cell(row=row, column=1, value="Prosečna starost reklamacije (dani)")
    ws.cell(row=row, column=2, value=df_predlog["Starost (dani)"].mean().round(0))
    row += 1
    
    ws.cell(row=row, column=1, value="Prosečno vreme u magacinu (dani)")
    ws.cell(row=row, column=2, value=df_predlog["Vreme u magacinu (dani)"].mean().round(0))
    row += 1
    
    ws.cell(row=row, column=1, value="Prosečna ukupna starost (dani)")
    ws.cell(row=row, column=2, value=df_predlog["Ukupna starost (dani)"].mean().round(0))
    
    for col in range(1, 14):
        ws.column_dimensions[get_column_letter(col)].width = 20

def upisi_metrike(ws, df):
    ws.cell(row=1, column=1, value="Napredne metrike")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=HEADER_BG)
    
    row = 3
    
    if "Nabavna cena" in df.columns and "Količina" in df.columns:
        df_copy = df.copy()
        if "Datum obrade" in df_copy.columns:
            df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
            danas = datetime.now()
            df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
        
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        pareto = df_copy.sort_values("Ukupna vrednost", ascending=False)
        pareto["Kumulativni procenat"] = (pareto["Ukupna vrednost"].cumsum() / pareto["Ukupna vrednost"].sum() * 100).round(2)
        
        ws.cell(row=row, column=1, value="PARETO ANALIZA (80/20)")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Redni broj", "Broj reklamacije", "Naziv artikla", "Vrednost (RSD)", "Kumulativni %"]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        primeni_stil_header(ws, len(headers))
        row += 1
        
        for idx, (_, r) in enumerate(pareto.head(20).iterrows(), 1):
            ws.cell(row=row, column=1).value = idx
            ws.cell(row=row, column=2).value = excel_compatible(r["Broj reklamacije"])
            ws.cell(row=row, column=3).value = excel_compatible(r["Naziv artikla"])
            ws.cell(row=row, column=4).value = excel_compatible(r["Ukupna vrednost"])
            ws.cell(row=row, column=5).value = excel_compatible(r["Kumulativni procenat"])
            row += 1
        
        granica = pareto[pareto["Kumulativni procenat"] <= 80].shape[0]
        ws.cell(row=row + 1, column=1, value=f"📊 {granica} artikala čini 80% ukupne vrednosti")
        ws.cell(row=row + 1, column=1).font = Font(bold=True, size=11, color="0066CC")
        row += 3
    
    if "Datum obrade" in df.columns:
        df_datum = df.copy()
        df_datum["Datum obrade"] = pd.to_datetime(df_datum["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_datum = df_datum[(df_datum["Datum obrade"].notna()) & (df_datum["Datum obrade"] <= danas)]
        
        if len(df_datum) > 0:
            df_datum["Starost (dani)"] = (danas - df_datum["Datum obrade"]).dt.days
            
            ws.cell(row=row, column=1, value="STAROST ARTIKALA")
            ws.cell(row=row, column=1).font = Font(bold=True, size=12)
            row += 1
            
            ws.cell(row=row, column=1, value="Prosečna starost (dani)")
            ws.cell(row=row, column=2, value=excel_compatible(df_datum["Starost (dani)"].mean().round(0)))
            row += 1
            
            ws.cell(row=row, column=1, value="Najstariji artikal (dani)")
            ws.cell(row=row, column=2, value=excel_compatible(df_datum["Starost (dani)"].max()))
            row += 1
            
            ws.cell(row=row, column=1, value="Najmlađi artikal (dani)")
            ws.cell(row=row, column=2, value=excel_compatible(df_datum["Starost (dani)"].min()))
            row += 2
            
            ws.cell(row=row, column=1, value="Top 10 najstarijih artikala")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11)
            row += 1
            
            headers = ["Naziv artikla", "Starost (dani)", "Broj reklamacije"]
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col_idx, value=h)
            primeni_stil_header(ws, len(headers))
            row += 1
            
            najstariji = df_datum.nlargest(10, "Starost (dani)")[["Naziv artikla", "Starost (dani)", "Broj reklamacije"]]
            for _, r in najstariji.iterrows():
                ws.cell(row=row, column=1).value = excel_compatible(r["Naziv artikla"])
                ws.cell(row=row, column=2).value = excel_compatible(r["Starost (dani)"])
                ws.cell(row=row, column=3).value = excel_compatible(r["Broj reklamacije"])
                row += 1
            row += 2
    
    if "Nabavna cena" in df.columns:
        ws.cell(row=row, column=1, value="TOP 10 NAJSKUPLJIH ARTIKALA")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        headers = ["Broj reklamacije", "Naziv artikla", "Nabavna cena (RSD)", "Brend"]
        for col_idx, h in enumerate(headers, start=1):
            ws.cell(row=row, column=col_idx, value=h)
        primeni_stil_header(ws, len(headers))
        row += 1
        
        najskuplji = df.nlargest(10, "Nabavna cena")[["Broj reklamacije", "Naziv artikla", "Nabavna cena", "Brend"]]
        for _, r in najskuplji.iterrows():
            ws.cell(row=row, column=1).value = excel_compatible(r["Broj reklamacije"])
            ws.cell(row=row, column=2).value = excel_compatible(r["Naziv artikla"])
            ws.cell(row=row, column=3).value = excel_compatible(r["Nabavna cena"])
            ws.cell(row=row, column=4).value = excel_compatible(r["Brend"])
            row += 1
        row += 2
    
    if "Serijski broj" in df.columns:
        duplikati = df[df["Serijski broj"].duplicated(keep=False)]
        ws.cell(row=row, column=1, value="DETEKCIJA DUPLIKATA")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        if len(duplikati) > 0:
            ws.cell(row=row, column=1, value=f"⚠️ Pronađeno {len(duplikati)} duplih serijskih brojeva")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="CC0000")
            row += 1
            
            headers = ["Serijski broj", "Naziv artikla", "Broj pojavljivanja", "Brojevi reklamacija"]
            for col_idx, h in enumerate(headers, start=1):
                ws.cell(row=row, column=col_idx, value=h)
            primeni_stil_header(ws, len(headers))
            row += 1
            
            for serijski, group in duplikati.groupby("Serijski broj"):
                ws.cell(row=row, column=1).value = excel_compatible(serijski)
                ws.cell(row=row, column=2).value = excel_compatible(group["Naziv artikla"].iloc[0])
                ws.cell(row=row, column=3).value = len(group)
                brojevi = ", ".join([str(b) for b in group["Broj reklamacije"].tolist()])
                ws.cell(row=row, column=4).value = brojevi
                row += 1
        else:
            ws.cell(row=row, column=1, value="✅ Nema duplih serijskih brojeva")
            ws.cell(row=row, column=1).font = Font(bold=True, size=11, color="008800")
        row += 2
    
    if "Brend" in df.columns and "Nabavna cena" in df.columns:
        ws.cell(row=row, column=1, value="PROSEČNA VREDNOST PO BRENDU")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1
        
        prosek_brend = df.groupby("Brend")["Nabavna cena"].mean().sort_values(ascending=False).head(10).reset_index()
        prosek_brend.columns = ["Brend", "Prosečna cena (RSD)"]
        
        ws.cell(row=row, column=1, value="Brend")
        ws.cell(row=row, column=2, value="Prosečna cena (RSD)")
        primeni_stil_header(ws, 2)
        row += 1
        
        for _, r in prosek_brend.iterrows():
            ws.cell(row=row, column=1).value = excel_compatible(r["Brend"])
            ws.cell(row=row, column=2).value = excel_compatible(r["Prosečna cena (RSD)"])
            row += 1
    
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 30

def formatiraj_i_sacuvaj(df):
    wb = Workbook()
    
    ws = wb.active
    ws.title = "zbirno"
    upisi_zbirno(ws, df)
    
    ws = wb.create_sheet("KPI")
    upisi_kpi(ws, df)
    
    ws = wb.create_sheet("Po klasifikaciji")
    upisi_grupisano(ws, df, "Klasifikacija reciklaže", "Klasifikacija reciklaže")
    
    ws = wb.create_sheet("Po šteti")
    upisi_grupisano(ws, df, "Klasifikacija štete", "Klasifikacija štete")
    
    ws = wb.create_sheet("Top brendovi")
    upisi_top_brendove(ws, df)
    
    ws = wb.create_sheet("Top grupe")
    upisi_top_grupe(ws, df)
    
    ws = wb.create_sheet("Vremenski trend")
    upisi_vremenski_trend(ws, df)
    
    ws = wb.create_sheet("Po prevozniku")
    upisi_prevoznike(ws, df)
    
    ws = wb.create_sheet("Analiza rokova")
    upisi_analiza_rokova(ws, df)
    
    ws = wb.create_sheet("Predlog prodaje 3. lice")
    upisi_predlog_prodaje(ws, df, st.session_state.kurs_evra, st.session_state.komada_po_paleti, st.session_state.cena_po_paleti_evri)
    
    ws = wb.create_sheet("Napredne metrike")
    upisi_metrike(ws, df)
    
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()

# ============================
# PDF GENERATOR (skraćen)
# ============================

def generisi_pdf(df):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=1.5*cm, leftMargin=1.5*cm, topMargin=1.5*cm, bottomMargin=1.5*cm)
    styles = getSampleStyleSheet()
    elements = []
    
    naslov_style = ParagraphStyle('Naslov', parent=styles['Heading1'], fontName=FONT_NAME, fontSize=24, textColor=colors.HexColor('#1F4E79'), alignment=1, spaceAfter=20, spaceBefore=30)
    sekcija_style = ParagraphStyle('Sekcija', parent=styles['Heading3'], fontName=FONT_NAME, fontSize=14, textColor=colors.HexColor('#2E75B6'), spaceAfter=8, spaceBefore=12)
    normal_style = ParagraphStyle('Normal', parent=styles['Normal'], fontName=FONT_NAME, fontSize=10, spaceAfter=4)
    bold_style = ParagraphStyle('Bold', parent=styles['Normal'], fontName=FONT_NAME, fontSize=10, textColor=colors.HexColor('#1F4E79'), spaceAfter=4)
    
    df_copy = df.copy()
    if "Datum obrade" in df_copy.columns:
        df_copy["Datum obrade"] = pd.to_datetime(df_copy["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_copy = df_copy[(df_copy["Datum obrade"].notna()) & (df_copy["Datum obrade"] <= danas)]
    
    ukupno = df_copy["Količina"].sum() if "Količina" in df_copy.columns else 0
    vrednost = (df_copy["Nabavna cena"] * df_copy["Količina"]).sum() if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns else 0
    broj_brendova = df_copy["Brend"].nunique() if "Brend" in df_copy.columns else 0
    
    # NASLOVNA
    elements.append(Spacer(1, 4*cm))
    elements.append(Paragraph("Tehnomanija", naslov_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph("IZVEŠTAJ O RECIKLAŽI", naslov_style))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"Generisano: {datetime.now().strftime('%d.%m.%Y %H:%M')}", normal_style))
    elements.append(Spacer(1, 2*cm))
    
    kpi_data = [["Ukupno artikala", f"{ukupno:,}"], ["Ukupna vrednost", f"{vrednost:,.2f} RSD"], ["Broj brendova", str(broj_brendova)]]
    kpi_table = Table(kpi_data, colWidths=[7*cm, 5*cm])
    kpi_table.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 11), ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (0,-1), colors.white), ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#CCCCCC')), ('ALIGN', (1,0), (1,-1), 'RIGHT')]))
    elements.append(kpi_table)
    elements.append(PageBreak())
    
    # 1. KLASIFIKACIJA RECIKLAŽE
    elements.append(Paragraph("1. Klasifikacija reciklaže", sekcija_style))
    if "Klasifikacija reciklaže" in df_copy.columns:
        klas = df_copy["Klasifikacija reciklaže"].value_counts().head(10).reset_index()
        klas.columns = ["Klasifikacija", "Broj"]
        if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
            df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
            vrednost_klas = df_copy.groupby("Klasifikacija reciklaže")["Ukupna vrednost"].sum().reset_index()
            vrednost_klas.columns = ["Klasifikacija", "Vrednost (RSD)"]
            klas = klas.merge(vrednost_klas, on="Klasifikacija", how="left")
        
        table_data = [["Klasifikacija", "Broj", "Vrednost (RSD)"]] + [[str(r["Klasifikacija"])[:40], str(r["Broj"]), f"{r['Vrednost (RSD)']:,.2f}" if 'Vrednost (RSD)' in r else "0"] for _, r in klas.iterrows()]
        t = Table(table_data, colWidths=[7*cm, 3*cm, 4*cm])
        t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 8), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (2,1), (2,-1), 'RIGHT')]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))
        
        fig, ax = plt.subplots(figsize=(6, 4))
        klas_pie = df_copy["Klasifikacija reciklaže"].value_counts().head(6)
        ax.pie(klas_pie.values, labels=klas_pie.index, autopct='%1.0f%%', startangle=90)
        ax.set_title('Klasifikacija reciklaže')
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        elements.append(Image(img_buffer, width=14*cm, height=8*cm))
    elements.append(PageBreak())
    
    # 2. KLASIFIKACIJA ŠTETE
    elements.append(Paragraph("2. Klasifikacija štete", sekcija_style))
    if "Klasifikacija štete" in df_copy.columns:
        steta = df_copy["Klasifikacija štete"].value_counts().head(10).reset_index()
        steta.columns = ["Klasifikacija štete", "Broj"]
        if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
            vrednost_steta = df_copy.groupby("Klasifikacija štete")["Ukupna vrednost"].sum().reset_index()
            vrednost_steta.columns = ["Klasifikacija štete", "Vrednost (RSD)"]
            steta = steta.merge(vrednost_steta, on="Klasifikacija štete", how="left")
        
        table_data = [["Klasifikacija štete", "Broj", "Vrednost (RSD)"]] + [[str(r["Klasifikacija štete"])[:40], str(r["Broj"]), f"{r['Vrednost (RSD)']:,.2f}" if 'Vrednost (RSD)' in r else "0"] for _, r in steta.iterrows()]
        t = Table(table_data, colWidths=[7*cm, 3*cm, 4*cm])
        t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 8), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (2,1), (2,-1), 'RIGHT')]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))
        
        fig, ax = plt.subplots(figsize=(6, 4))
        steta_bar = df_copy["Klasifikacija štete"].value_counts().head(6)
        ax.barh(steta_bar.index, steta_bar.values, color='#1F4E79')
        ax.set_xlabel('Broj')
        ax.set_title('Klasifikacija štete')
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        elements.append(Image(img_buffer, width=14*cm, height=8*cm))
    elements.append(PageBreak())
    
    # 3. TOP 10 BRENDOVA
    elements.append(Paragraph("3. Top 10 brendova", sekcija_style))
    if "Brend" in df_copy.columns:
        if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
            brend_analiza = df_copy.groupby("Brend").agg({
                "Broj reklamacije": "count",
                "Ukupna vrednost": "sum"
            }).reset_index()
            brend_analiza.columns = ["Brend", "Broj artikala", "Ukupna vrednost (RSD)"]
        else:
            brend_analiza = df_copy["Brend"].value_counts().head(10).reset_index()
            brend_analiza.columns = ["Brend", "Broj artikala"]
            brend_analiza["Ukupna vrednost (RSD)"] = 0
        
        brend_analiza = brend_analiza.sort_values("Broj artikala", ascending=False).head(10)
        
        table_data = [["Brend", "Broj", "Vrednost (RSD)"]] + [[str(r["Brend"]), str(r["Broj artikala"]), f"{r['Ukupna vrednost (RSD)']:,.2f}"] for _, r in brend_analiza.iterrows()]
        t = Table(table_data, colWidths=[7*cm, 3*cm, 4*cm])
        t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 8), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (2,1), (2,-1), 'RIGHT')]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))
        
        fig, ax = plt.subplots(figsize=(6, 4))
        brend_bar = df_copy["Brend"].value_counts().head(10)
        ax.bar(brend_bar.index, brend_bar.values, color='#1F4E79')
        ax.set_ylabel('Broj')
        ax.set_title('Top 10 brendova')
        plt.xticks(rotation=45, ha='right')
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        elements.append(Image(img_buffer, width=14*cm, height=8*cm))
    elements.append(PageBreak())
    
    # 4. VREMENSKI TREND
    elements.append(Paragraph("4. Vremenski trend", sekcija_style))
    if "Datum obrade" in df_copy.columns:
        df_datum = df_copy.copy()
        df_datum["Datum obrade"] = pd.to_datetime(df_datum["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_datum = df_datum[(df_datum["Datum obrade"].notna()) & (df_datum["Datum obrade"] <= danas) & (df_datum["Datum obrade"] >= datetime(2020, 1, 1))]
        
        if len(df_datum) > 0:
            df_datum["Mesec"] = df_datum["Datum obrade"].dt.to_period("M")
            trend = df_datum.groupby("Mesec").agg({"Broj reklamacije": "count"}).reset_index()
            trend.columns = ["Mesec", "Broj artikala"]
            trend["Mesec"] = trend["Mesec"].astype(str)
            
            if "Ukupna vrednost" in df_datum.columns:
                vrednost_trend = df_datum.groupby("Mesec")["Ukupna vrednost"].sum().reset_index()
                vrednost_trend.columns = ["Mesec", "Vrednost (RSD)"]
                vrednost_trend["Mesec"] = vrednost_trend["Mesec"].astype(str)
                trend = trend.merge(vrednost_trend, on="Mesec", how="left")
            
            table_data = [["Mesec", "Broj", "Vrednost (RSD)"]] + [[str(r["Mesec"]), str(r["Broj artikala"]), f"{r['Vrednost (RSD)']:,.2f}" if 'Vrednost (RSD)' in r else "0"] for _, r in trend.iterrows()]
            t = Table(table_data, colWidths=[5*cm, 3*cm, 5*cm])
            t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 8), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (2,1), (2,-1), 'RIGHT')]))
            elements.append(t)
            elements.append(Spacer(1, 0.5*cm))
            
            fig, ax = plt.subplots(figsize=(6, 4))
            ax.plot(trend["Mesec"], trend["Broj artikala"], marker='o', color='#1F4E79')
            ax.set_ylabel('Broj')
            ax.set_title('Vremenski trend')
            ax.grid(True, linestyle='--', alpha=0.5)
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            elements.append(Image(img_buffer, width=14*cm, height=8*cm))
        else:
            elements.append(Paragraph("Nema validnih datuma za prikaz", normal_style))
    elements.append(PageBreak())
    
    # 5. PARETO ANALIZA
    elements.append(Paragraph("5. Pareto analiza (80/20)", sekcija_style))
    if "Nabavna cena" in df_copy.columns and "Količina" in df_copy.columns:
        df_copy["Ukupna vrednost"] = df_copy["Nabavna cena"] * df_copy["Količina"]
        pareto = df_copy.sort_values("Ukupna vrednost", ascending=False)
        pareto["Kumulativni %"] = (pareto["Ukupna vrednost"].cumsum() / pareto["Ukupna vrednost"].sum() * 100).round(2)
        
        table_data = [["Broj reklamacije", "Naziv", "Vrednost", "Kum.%"]]
        for _, r in pareto.head(15).iterrows():
            table_data.append([str(r["Broj reklamacije"])[:12], str(r["Naziv artikla"])[:25], f"{r['Ukupna vrednost']:,.0f}", f"{r['Kumulativni %']}%"])
        t = Table(table_data, colWidths=[2.5*cm, 5*cm, 3*cm, 2.5*cm])
        t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 7), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC'))]))
        elements.append(t)
        elements.append(Spacer(1, 0.3*cm))
        granica = pareto[pareto["Kumulativni %"] <= 80].shape[0]
        elements.append(Paragraph(f"📊 {granica} artikala čini 80% ukupne vrednosti", bold_style))
        elements.append(Spacer(1, 0.5*cm))
        
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.bar(range(20), pareto["Ukupna vrednost"].head(20), color='#1F4E79')
        ax.axhline(y=pareto["Ukupna vrednost"].sum() * 0.8, color='red', linestyle='--', label='80% granica')
        ax.set_xlabel('Redni broj')
        ax.set_ylabel('Vrednost')
        ax.set_title('Pareto analiza')
        ax.legend()
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        elements.append(Image(img_buffer, width=14*cm, height=8*cm))
    elements.append(PageBreak())
    
    # 6. TOP 10 NAJSKUPLJIH
    elements.append(Paragraph("6. Top 10 najskupljih artikala", sekcija_style))
    if "Nabavna cena" in df_copy.columns:
        naj = df_copy.nlargest(10, "Nabavna cena")[["Broj reklamacije", "Naziv artikla", "Nabavna cena", "Brend"]]
        table_data = [["Broj reklamacije", "Naziv", "Cena (RSD)", "Brend"]]
        for _, r in naj.iterrows():
            table_data.append([str(r["Broj reklamacije"])[:12], str(r["Naziv artikla"])[:25], f"{r['Nabavna cena']:,.2f}", str(r["Brend"])])
        t = Table(table_data, colWidths=[2.5*cm, 5*cm, 3*cm, 2.5*cm])
        t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 7), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (2,1), (2,-1), 'RIGHT')]))
        elements.append(t)
        elements.append(Spacer(1, 0.5*cm))
        
        fig, ax = plt.subplots(figsize=(6, 4))
        ax.barh(naj["Naziv artikla"].str[:20], naj["Nabavna cena"], color='#1F4E79')
        ax.set_xlabel('Cena (RSD)')
        ax.set_title('Top 10 najskupljih')
        img_buffer = io.BytesIO()
        plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
        img_buffer.seek(0)
        plt.close()
        elements.append(Image(img_buffer, width=14*cm, height=8*cm))
    elements.append(PageBreak())
    
    # 7. ANALIZA ROKOVA
    elements.append(Paragraph("7. Analiza rokova (30 dana)", sekcija_style))
    if "Status roka" in df_copy.columns and "Starost (dani)" in df_copy.columns:
        df_rok = df_copy.copy()
        df_rok["Datum obrade"] = pd.to_datetime(df_rok["Datum obrade"], errors="coerce")
        danas = datetime.now()
        df_rok = df_rok[(df_rok["Datum obrade"].notna()) & (df_rok["Datum obrade"] <= danas)]
        
        if len(df_rok) > 0:
            u_roku = df_rok[df_rok["Status roka"] == "✅ U roku"].shape[0]
            prekoraceno = df_rok[df_rok["Status roka"] == "❌ Prekoračen"].shape[0]
            ukupno_rok = len(df_rok)
            
            table_data = [
                ["✅ U roku (≤30 dana)", u_roku, f"{u_roku/ukupno_rok*100:.1f}%" if ukupno_rok > 0 else "0%"],
                ["❌ Prekoračen (>30 dana)", prekoraceno, f"{prekoraceno/ukupno_rok*100:.1f}%" if ukupno_rok > 0 else "0%"],
                ["Prosečno prekoračenje", f"{df_rok['Prekoračenje (dani)'].mean().round(0):.0f} dana", ""],
                ["Maksimalno prekoračenje", f"{df_rok['Prekoračenje (dani)'].max():.0f} dana", ""]
            ]
            t = Table(table_data, colWidths=[6*cm, 4*cm, 4*cm])
            t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 9), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC')), ('ALIGN', (1,0), (2,-1), 'CENTER')]))
            elements.append(t)
            elements.append(Spacer(1, 0.5*cm))
            
            elements.append(Paragraph("Top 10 najvećih prekoračenja", sekcija_style))
            table_data = [["Broj reklamacije", "Naziv", "Starost", "Prekoračenje", "Otvorio"]]
            for _, r in df_rok.nlargest(10, "Prekoračenje (dani)").iterrows():
                table_data.append([
                    str(r["Broj reklamacije"])[:12],
                    str(r["Naziv artikla"])[:25],
                    str(r["Starost (dani)"]),
                    str(r["Prekoračenje (dani)"]),
                    str(r["reklamacija otvorena od strane"])[:15]
                ])
            t = Table(table_data, colWidths=[2.5*cm, 5*cm, 2*cm, 2.5*cm, 3*cm])
            t.setStyle(TableStyle([('FONTNAME', (0,0), (-1,-1), FONT_NAME), ('FONTSIZE', (0,0), (-1,-1), 7), ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1F4E79')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CCCCCC'))]))
            elements.append(t)
            elements.append(Spacer(1, 0.5*cm))
            
            fig, ax = plt.subplots(figsize=(6, 4))
            statusi = ["✅ U roku", "❌ Prekoračen"]
            brojevi = [u_roku, prekoraceno]
            colors_pie = ['#2E8B57', '#DC143C']
            ax.pie(brojevi, labels=statusi, autopct='%1.0f%%', colors=colors_pie, startangle=90)
            ax.set_title('Status rokova reklamacija')
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=100, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            elements.append(Image(img_buffer, width=10*cm, height=8*cm))
        else:
            elements.append(Paragraph("Nema validnih podataka za analizu rokova", normal_style))
    else:
        elements.append(Paragraph("Nedostaju podaci za analizu rokova", normal_style))
    
    elements.append(PageBreak())
    elements.append(Spacer(1, 5*cm))
    elements.append(Paragraph("Izveštaj generisan automatski", normal_style))
    elements.append(Paragraph(f"Tehnomanija © {datetime.now().year}", normal_style))
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

# ============================
# EMAIL SLANJE
# ============================

def posalji_email(primalac, excel_data=None, pdf_data=None, dodatne_adrese=None):
    try:
        sender = st.secrets["email"]["sender"]
        password = st.secrets["email"]["password"]
        smtp_server = st.secrets["email"]["smtp_server"]
        smtp_port = st.secrets["email"]["smtp_port"]
        
        msg = MIMEMultipart()
        msg["From"] = sender
        msg["To"] = primalac
        msg["Subject"] = f"Izveštaj o reciklaži – {datetime.now().strftime('%d.%m.%Y')}"
        
        if dodatne_adrese:
            msg["CC"] = ", ".join(dodatne_adrese)
        
        body = f"Poštovani,\n\nU prilogu vam dostavljamo izveštaj o reciklaži sa stanjem od {datetime.now().strftime('%d.%m.%Y')}.\n\nIzveštaj sadrži analizu rokova (30 dana) i predlog za prodaju trećem licu.\n\nIzveštaj je generisan automatski.\n\nS poštovanjem,\nTehnomanija"
        msg.attach(MIMEText(body, "plain"))
        
        if excel_data is not None and isinstance(excel_data, bytes):
            excel_part = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            excel_part.set_payload(excel_data)
            encoders.encode_base64(excel_part)
            excel_part.add_header('Content-Disposition', f'attachment; filename=izvestaj_reciklaza_{datetime.now().strftime("%Y%m%d")}.xlsx')
            msg.attach(excel_part)
        
        if pdf_data is not None and isinstance(pdf_data, bytes):
            pdf_part = MIMEBase('application', 'pdf')
            pdf_part.set_payload(pdf_data)
            encoders.encode_base64(pdf_part)
            pdf_part.add_header('Content-Disposition', f'attachment; filename=izvestaj_reciklaza_{datetime.now().strftime("%Y%m%d")}.pdf')
            msg.attach(pdf_part)
        
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()
        server.login(sender, password)
        
        svi_primaoci = [primalac]
        if dodatne_adrese:
            svi_primaoci.extend(dodatne_adrese)
        
        server.send_message(msg)
        server.quit()
        return True, "Email je uspešno poslat!"
    
    except Exception as e:
        return False, f"Greška pri slanju emaila: {str(e)}"

# ============================
# STREAMLIT UI
# ============================

st.title("🔄 Izveštaj o reciklaži")
uploaded_file = st.file_uploader("Izaberi Excel fajl", type=["xlsx"])

if uploaded_file is not None:
    with st.spinner("Učitavam..."):
        reciklaza, reklamacije = ucitaj_podatke(uploaded_file)
    if reciklaza is not None and reklamacije is not None:
        with st.spinner("Spajam..."):
            lookup = pripremi_lookup(reklamacije)
            df = spoji(reciklaza, lookup)
        st.success(f"✅ Učitano: {len(reciklaza)} redova, spojeno: {len(df)} redova")
        
        # Debug: Provera novembra
        if "Datum obrade" in df.columns:
            df["Datum obrade"] = pd.to_datetime(df["Datum obrade"], errors="coerce")
            novembar = df[df["Datum obrade"].dt.month == 11]
            if len(novembar) > 0:
                st.warning(f"⚠️ Pronađeno {len(novembar)} redova sa novembrom!")
            else:
                st.success("✅ Nema novembarskih datuma!")
        
        st.markdown("---")
        
        tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
            "📊 Pregled",
            "📈 Napredne metrike",
            "🏷️ Analize",
            "📋 Tabela",
            "📤 Izvoz",
            "⏰ Rokovi",
            "📦 Prodaja 3. licu"
        ])
        
        with tab1:
            st.subheader("Ključni pokazatelji")
            col1, col2, col3, col4 = st.columns(4)
            ukupno = df["Količina"].sum() if "Količina" in df.columns else 0
            vrednost = (df["Nabavna cena"] * df["Količina"]).sum() if "Nabavna cena" in df.columns and "Količina" in df.columns else 0
            col1.metric("📦 Ukupno artikala", f"{ukupno:,}")
            col2.metric("💰 Ukupna vrednost", f"{vrednost:,.2f} RSD")
            col3.metric("🏷️ Broj brendova", df["Brend"].nunique() if "Brend" in df.columns else 0)
            col4.metric("📂 Robnih grupa", df["Robna grupa"].nunique() if "Robna grupa" in df.columns else 0)
            
            st.markdown("---")
            if "Status roka" in df.columns:
                u_roku = df[df["Status roka"] == "✅ U roku"].shape[0]
                prekoraceno = df[df["Status roka"] == "❌ Prekoračen"].shape[0]
                col1, col2, col3 = st.columns(3)
                col1.metric("✅ U roku (≤30 dana)", u_roku)
                col2.metric("❌ Prekoračen (>30 dana)", prekoraceno)
                if u_roku + prekoraceno > 0:
                    col3.metric("📊 Procenat u roku", f"{u_roku/(u_roku+prekoraceno)*100:.1f}%")
        
        with tab2:
            st.subheader("Napredne metrike")
            if "Nabavna cena" in df.columns and "Količina" in df.columns:
                df["Ukupna vrednost"] = df["Nabavna cena"] * df["Količina"]
                pareto = df.sort_values("Ukupna vrednost", ascending=False)
                pareto["Kumulativni %"] = (pareto["Ukupna vrednost"].cumsum() / pareto["Ukupna vrednost"].sum() * 100).round(2)
                fig = px.bar(pareto.head(20), x="Naziv artikla", y="Ukupna vrednost", title="Pareto analiza - Top 20 artikala")
                st.plotly_chart(fig, use_container_width=True)
                st.info(f"📊 {pareto[pareto['Kumulativni %'] <= 80].shape[0]} artikala čini 80% ukupne vrednosti")
            
            st.markdown("---")
            st.subheader("🕐 Starost reklamacija")
            if "datum otvaranja" in df.columns and "Datum obrade" in df.columns:
                df_starost = df.copy()
                df_starost["datum otvaranja"] = pd.to_datetime(df_starost["datum otvaranja"], errors="coerce")
                df_starost["Datum obrade"] = pd.to_datetime(df_starost["Datum obrade"], errors="coerce")
                danas = datetime.now()
                df_starost = df_starost[
                    (df_starost["Datum obrade"].notna()) & 
                    (df_starost["datum otvaranja"].notna()) &
                    (df_starost["Datum obrade"] <= danas)
                ]
                if len(df_starost) > 0:
                    df_starost["Starost (dani)"] = (df_starost["Datum obrade"] - df_starost["datum otvaranja"]).dt.days
                    col1, col2, col3 = st.columns(3)
                    col1.metric("📅 Prosečna starost", f"{df_starost['Starost (dani)'].mean().round(0):.0f} dana")
                    col2.metric("⏳ Najstarija reklamacija", f"{df_starost['Starost (dani)'].max():.0f} dana")
                    col3.metric("🆕 Najmlađa reklamacija", f"{df_starost['Starost (dani)'].min():.0f} dana")
                    fig = px.histogram(df_starost, x="Starost (dani)", nbins=20, title="Distribucija starosti reklamacija")
                    st.plotly_chart(fig, use_container_width=True)
        
        with tab3:
            st.subheader("Sve analize")
            if "Klasifikacija reciklaže" in df.columns:
                klas = df["Klasifikacija reciklaže"].value_counts().reset_index()
                klas.columns = ["Klasifikacija", "Broj"]
                st.dataframe(klas, use_container_width=True)
        
        with tab4:
            st.dataframe(df, use_container_width=True)
        
        with tab5:
            st.subheader("📤 Izvoz izveštaja")
            
            excel_data = None
            pdf_data = None
            
            col1, col2 = st.columns(2)
            with col1:
                try:
                    excel_data = formatiraj_i_sacuvaj(df)
                    st.download_button(
                        label="📥 Preuzmi Excel (11 sheet-ova)",
                        data=excel_data,
                        file_name=f"izvestaj_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Greška: {e}")
            
            with col2:
                with st.spinner("Generišem PDF..."):
                    try:
                        pdf_data = generisi_pdf(df)
                        st.download_button(
                            label="📄 Preuzmi PDF (sa analizom rokova)",
                            data=pdf_data,
                            file_name=f"izvestaj_{datetime.now().strftime('%Y%m%d')}.pdf",
                            mime="application/pdf",
                            use_container_width=True
                        )
                    except Exception as e:
                        st.error(f"Greška: {e}")
            
            st.markdown("---")
            st.markdown("### 📧 Pošalji izveštaj emailom")
            with st.form("email_form"):
                col1, col2 = st.columns(2)
                with col1:
                    email_primalac = st.text_input("Email adresa (obavezno)", placeholder="office@tehnomanija.rs")
                with col2:
                    email_cc = st.text_input("Dodatne adrese (CC, odvojene zarezom)", placeholder="nikola@tehnomanija.rs, milica@tehnomanija.rs")
                
                col1, col2, col3 = st.columns([1, 1, 2])
                with col1:
                    posalji_pdf = st.checkbox("📄 Dodaj PDF", value=True)
                with col2:
                    poslati_excel = st.checkbox("📊 Dodaj Excel", value=True)
                
                submitted = st.form_submit_button("📩 Pošalji")
                
                if submitted:
                    if not email_primalac:
                        st.error("Unesite email adresu")
                    else:
                        cc_list = [email.strip() for email in email_cc.split(",") if email.strip()] if email_cc else []
                        excel_attach = excel_data if (poslati_excel and excel_data is not None) else None
                        pdf_attach = pdf_data if (posalji_pdf and pdf_data is not None) else None
                        
                        if not excel_attach and not pdf_attach:
                            st.warning("Nijedan prilog nije odabran")
                        else:
                            with st.spinner("Šaljem..."):
                                success, message = posalji_email(email_primalac, excel_attach, pdf_attach, cc_list)
                                if success:
                                    st.success(f"✅ {message}")
                                else:
                                    st.error(f"❌ {message}")
        
        with tab6:
            st.subheader("⏰ Analiza rokova (30 dana)")
            if "Status roka" in df.columns:
                df_rok = df.copy()
                if "Datum obrade" in df_rok.columns:
                    df_rok["Datum obrade"] = pd.to_datetime(df_rok["Datum obrade"], errors="coerce")
                    danas = datetime.now()
                    df_rok = df_rok[(df_rok["Datum obrade"].notna()) & (df_rok["Datum obrade"] <= danas)]
                
                if len(df_rok) > 0:
                    u_roku = df_rok[df_rok["Status roka"] == "✅ U roku"].shape[0]
                    prekoraceno = df_rok[df_rok["Status roka"] == "❌ Prekoračen"].shape[0]
                    ukupno_rok = len(df_rok)
                    
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("📋 Ukupno", ukupno_rok)
                    col2.metric("✅ U roku", u_roku)
                    col3.metric("❌ Prekoračen", prekoraceno)
                    col4.metric("📊 Procenat u roku", f"{u_roku/ukupno_rok*100:.1f}%" if ukupno_rok > 0 else "0%")
                    
                    st.markdown("---")
                    col1, col2 = st.columns(2)
                    with col1:
                        st.subheader("Prosečno prekoračenje")
                        st.metric("Dani", f"{df_rok['Prekoračenje (dani)'].mean().round(0):.0f} dana")
                        st.metric("Maksimalno", f"{df_rok['Prekoračenje (dani)'].max():.0f} dana")
                    
                    with col2:
                        fig = px.pie(
                            values=[u_roku, prekoraceno],
                            names=["✅ U roku", "❌ Prekoračen"],
                            title="Status rokova",
                            color_discrete_sequence=['#2E8B57', '#DC143C']
                        )
                        st.plotly_chart(fig, use_container_width=True)
                    
                    st.markdown("---")
                    st.subheader("Top 10 najvećih prekoračenja")
                    st.dataframe(
                        df_rok.nlargest(10, "Prekoračenje (dani)")[
                            ["Broj reklamacije", "Naziv artikla", "Starost (dani)", "Prekoračenje (dani)", "reklamacija otvorena od strane"]
                        ],
                        use_container_width=True
                    )
                    
                    st.markdown("---")
                    st.subheader("Filter po osobi koja je otvorila reklamaciju")
                    if "reklamacija otvorena od strane" in df_rok.columns:
                        osobe = ["Sve"] + sorted(df_rok["reklamacija otvorena od strane"].dropna().unique().tolist())
                        izabrana_osoba = st.selectbox("Izaberi osobu", osobe)
                        
                        if izabrana_osoba != "Sve":
                            df_filtered = df_rok[df_rok["reklamacija otvorena od strane"] == izabrana_osoba]
                        else:
                            df_filtered = df_rok
                        
                        st.dataframe(
                            df_filtered[["Broj reklamacije", "Naziv artikla", "Starost (dani)", "Prekoračenje (dani)", "Status roka"]],
                            use_container_width=True
                        )
                else:
                    st.info("Nema validnih podataka za analizu rokova")
            else:
                st.warning("Nedostaju podaci za analizu rokova. Proverite da li Excel fajl sadrži kolonu 'Reklamacija: Created Date'.")
        
        with tab7:
            st.subheader("📦 Predlog za prodaju trećem licu")
            st.info("""
            **Kriterijum za prodaju:**
            - Starost reklamacije > 30 dana (zakonski rok istekao)
            - Vreme u magacinu reciklaže > 30 dana (čeka na prodaju)
            - Ukupno > 60 dana od otvaranja reklamacije
            """)
            
            col1, col2, col3 = st.columns(3)
            with col1:
                kurs_evra = st.number_input("💶 Kurs evra (RSD)", min_value=50.0, max_value=200.0, value=st.session_state.kurs_evra, step=0.5)
                st.session_state.kurs_evra = kurs_evra
            with col2:
                komada_po_paleti = st.number_input("📦 Komada po paleti", min_value=1, max_value=500, value=st.session_state.komada_po_paleti, step=1)
                st.session_state.komada_po_paleti = komada_po_paleti
            with col3:
                cena_po_paleti = st.number_input("💰 Cena po paleti (€)", min_value=10.0, max_value=500.0, value=st.session_state.cena_po_paleti_evri, step=5.0)
                st.session_state.cena_po_paleti_evri = cena_po_paleti
            
            df_predlog, sazetak = generisi_predlog_prodaje(df, kurs_evra, komada_po_paleti, cena_po_paleti)
            
            if df_predlog is not None and len(df_predlog) > 0:
                st.markdown("### 📊 Sažetak predloga")
                col1, col2, col3, col4 = st.columns(4)
                col1.metric("📦 Ukupno artikala", sazetak["Ukupno artikala"])
                col2.metric("📦 Ukupno paleta", sazetak["Ukupno paleta"])
                col3.metric("💰 Ponuda (€)", f"{sazetak['Ukupna vrednost ponude (€)']:,.2f} €")
                col4.metric("📉 Gubitak (RSD)", f"{sazetak['Ukupni gubitak (RSD)']:,.2f} RSD")
                
                col1, col2, col3 = st.columns(3)
                col1.metric("⏳ Prosečna starost reklamacije", f"{sazetak['Prosečna starost reklamacije (dani)']:.0f} dana")
                col2.metric("🏠 Prosečno vreme u magacinu", f"{sazetak['Prosečno vreme u magacinu (dani)']:.0f} dana")
                col3.metric("📅 Prosečna ukupna starost", f"{sazetak['Prosečna ukupna starost (dani)']:.0f} dana")
                
                st.markdown("### 📋 Detaljna tabela")
                st.dataframe(df_predlog, use_container_width=True)
                
                st.markdown("### 📤 Izvezi predlog")
                try:
                    wb_predlog = Workbook()
                    ws_predlog = wb_predlog.active
                    ws_predlog.title = "Predlog prodaje"
                    upisi_predlog_prodaje(ws_predlog, df, kurs_evra, komada_po_paleti, cena_po_paleti)
                    output_predlog = io.BytesIO()
                    wb_predlog.save(output_predlog)
                    predlog_data = output_predlog.getvalue()
                    
                    st.download_button(
                        label="📥 Preuzmi predlog (Excel)",
                        data=predlog_data,
                        file_name=f"predlog_prodaje_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Greška: {e}")
            else:
                st.success("✅ Trenutno nema artikala koji ispunjavaju kriterijume za prodaju trećem licu.")
